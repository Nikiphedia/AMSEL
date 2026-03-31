#!/usr/bin/env python3
"""
Extrahiert mehrsprachige Vogelnamen aus der IOC Multilingual World Bird List (XLSX)
und schreibt sie als JSON-Zwischendatei fuer den Generator.

Input:  D:/80002/AMSEL/Multiling IOC 15.1_d.xlsx
Output: D:/80002/AMSEL/tools/data/ioc_names.json

Ausgabeformat:
{
  "Turdus_merula": {
    "order": "PASSERIFORMES",
    "family": "Turdidae",
    "de": "Amsel",
    "en": "Common Blackbird",
    "fr": "Merle noir",
    ...
  }
}
"""

import json
import sys
import os

# openpyxl muss installiert sein (pip install openpyxl)
from openpyxl import load_workbook

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'Multiling IOC 15.1_d.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'data', 'ioc_names.json')

# Spalten-Mapping: IOC-Spaltenname -> EU-Sprachcode
# Irisch (ga) und Maltesisch (mt) fehlen in der IOC-Datei
COLUMN_MAP = {
    'German': 'de',
    'English': 'en',
    'French': 'fr',
    'Italian': 'it',
    'Spanish': 'es',
    'Portuguese (Lusophone)': 'pt',
    'Dutch': 'nl',
    'Danish': 'da',
    'Swedish': 'sv',
    'Finnish': 'fi',
    'Polish': 'pl',
    'Czech': 'cs',
    'Slovak': 'sk',
    'Hungarian': 'hu',
    'Romanian': 'ro',
    'Bulgarian': 'bg',
    'Croatian': 'hr',
    'Slovenian': 'sl',
    'Estonian': 'et',
    'Latvian': 'lv',
    'Lithuanian': 'lt',
    'Greek': 'el',
    'Ukrainian': 'uk',
    # ga (Irisch) und mt (Maltesisch) nicht in IOC enthalten
}

# Zusaetzliche Metadaten-Spalten
META_COLUMNS = {
    'Order': 'order',
    'Family': 'family',
}

def main():
    print(f'Lade IOC Multilingual XLSX: {os.path.abspath(XLSX_PATH)}')
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['List']

    # Header-Zeile lesen (Zeile 1)
    headers = None
    result = {}
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(h) if h else '' for h in row]
            # Spaltenindizes fuer Sprachen ermitteln
            sci_col = headers.index('IOC_15.1')
            lang_cols = {}
            for ioc_name, lang_code in COLUMN_MAP.items():
                try:
                    lang_cols[lang_code] = headers.index(ioc_name)
                except ValueError:
                    print(f'  WARNUNG: Spalte "{ioc_name}" ({lang_code}) nicht gefunden')
            meta_cols = {}
            for ioc_name, key in META_COLUMNS.items():
                try:
                    meta_cols[key] = headers.index(ioc_name)
                except ValueError:
                    print(f'  WARNUNG: Spalte "{ioc_name}" nicht gefunden')
            print(f'  Gefundene Sprachen: {len(lang_cols)} von {len(COLUMN_MAP)}')
            continue

        # Datenzeilen
        sci_name = row[sci_col] if row[sci_col] else None
        if not sci_name:
            skipped += 1
            continue

        sci_name = str(sci_name).strip()
        if not sci_name or ' ' not in sci_name:
            skipped += 1
            continue

        # Unterstrich-Format fuer Key
        key = sci_name.replace(' ', '_')

        entry = {}
        # Metadaten
        for meta_key, col_idx in meta_cols.items():
            val = row[col_idx]
            entry[meta_key] = str(val).strip() if val else ''

        # Sprachnamen
        for lang_code, col_idx in lang_cols.items():
            val = row[col_idx]
            if val:
                name = str(val).strip()
                # Klammerzusaetze bei Griechisch etc. entfernen: "(Κοινός) Κότσυφας" -> "Κότσυφας"
                if name.startswith('(') and ') ' in name:
                    name = name[name.index(') ') + 2:]
                entry[lang_code] = name
            else:
                entry[lang_code] = ''

        result[key] = entry

    wb.close()

    # Ausgabe
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=1)

    print(f'\nErgebnis:')
    print(f'  Arten extrahiert: {len(result)}')
    print(f'  Zeilen uebersprungen: {skipped}')
    print(f'  Ausgabe: {os.path.abspath(OUTPUT_PATH)}')

    # Stichproben
    for test_sp in ['Turdus_merula', 'Parus_major', 'Falco_peregrinus']:
        if test_sp in result:
            names = result[test_sp]
            filled = sum(1 for k, v in names.items() if v and k not in ('order', 'family'))
            print(f'  {test_sp}: {filled} Sprachen (de={names.get("de", "?")})')

if __name__ == '__main__':
    main()
